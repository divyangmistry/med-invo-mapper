"""
excel_writer.py — Daily Excel file generation using openpyxl.

Behaviour
---------
- On first call of the day, creates `Inventory_YYYY-MM-DD.xlsx` with styled headers.
- On subsequent calls, appends a new data row.
- Never overwrites existing data; always appends.
"""
from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import Config
from extractor import ExtractionResult

logger = logging.getLogger(__name__)


# ── Column definitions (order matters) ───────────────────────────────────────

HEADERS = [
    "Timestamp",
    "Vendor Name",
    "Invoice Number",
    "Bill Date",
    "Medicine Name",
    "Medicine Code",
    "Batch Number",
    "Mfg. Date",
    "Exp. Date",
    "Quantity",
    "Unit",
    "Free Qty",
    "MRP",
    "PTR",
    "Disc %",
    "Disc Amt",
    "Base Amount",
    "GST %",
    "Amount",
    "HSN Code",
    "Location",
    "SGST",
    "CGST",
    "IGST",
    "Cess",
    "Total Amount",
    "Confidence",
    "Source Image",
    "Notes",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="EBF3FA")
NORMAL_FONT = Font(name="Calibri", size=10)
FLAG_COLORS = {
    "OK": "C6EFCE",
    "MANUAL_REVIEW": "FFEB9C",
    "RETRY": "FFC7CE",
}
COL_WIDTHS = [
    20,   # Timestamp
    30,   # Vendor Name
    20,   # Invoice Number
    14,   # Bill Date
    35,   # Medicine Name
    18,   # Medicine Code
    18,   # Batch Number
    14,   # Mfg. Date
    14,   # Exp. Date
    10,   # Quantity
    10,   # Unit
    10,   # Free Qty
    12,   # MRP
    12,   # PTR
    10,   # Disc %
    12,   # Disc Amt
    14,   # Base Amount
    10,   # GST %
    14,   # Amount
    14,   # HSN Code
    14,   # Location
    12,   # SGST
    12,   # CGST
    12,   # IGST
    12,   # Cess
    14,   # Total Amount
    18,   # Confidence
    35,   # Source Image
    25,   # Notes
]


def _get_daily_file_path(output_dir: Optional[Path] = None) -> Path:
    out = output_dir or Config.OUTPUT_DIR
    out.mkdir(parents=True, exist_ok=True)
    today = date.today().strftime("%Y-%m-%d")
    return out / f"Inventory_{today}.xlsx"


def _create_new_workbook(file_path: Path) -> openpyxl.Workbook:
    """Create a fresh workbook with styled headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Inventory"

    ws.append(HEADERS)
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[col_idx - 1]

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    wb.save(file_path)
    logger.info("Created new daily workbook: %s", file_path.name)
    return wb


def append_to_excel(
    result: ExtractionResult,
    timestamp: Optional[str] = None,
    source_image: Optional[str] = None,
    notes: Optional[str] = None,
    output_dir: Optional[Path] = None,
) -> Path:
    """
    Append multiple extraction items as new rows to today's Excel file.

    Parameters
    ----------
    result       : ExtractionResult  — parsed VLM result (contains items)
    timestamp    : str               — ISO datetime string (defaults to now)
    source_image : str               — filename of the processed image
    notes        : str               — optional manual notes
    output_dir   : Path              — override default output directory

    Returns
    -------
    Path — path to the Excel file that was written
    """
    from datetime import datetime, timezone

    file_path = _get_daily_file_path(output_dir)

    # Load or create workbook
    if file_path.exists():
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    else:
        wb = _create_new_workbook(file_path)
        ws = wb.active

    ts = timestamp or datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    
    if not result.items:
        logger.warning("No items to append to Excel for %s", source_image)
        return file_path

    for item in result.items:
        row_data = [
            ts,
            result.vendor_name,
            result.invoice_number,
            getattr(result, "bill_date", ""),
            item.medicine_name,
            item.medicine_code,
            item.batch_number,
            item.manufacturing_date,
            item.expiry_date,
            item.quantity,
            getattr(item, "unit", ""),
            getattr(item, "free_quantity", 0),
            getattr(item, "mrp", ""),
            getattr(item, "ptr", ""),
            getattr(item, "discount_percent", ""),
            getattr(item, "discount_amount", ""),
            getattr(item, "base_amount", ""),
            getattr(item, "gst_percent", ""),
            getattr(item, "amount", ""),
            getattr(item, "hsn_code", ""),
            getattr(item, "location", ""),
            getattr(result, "sgst_amount", ""),
            getattr(result, "cgst_amount", ""),
            getattr(result, "igst_amount", ""),
            getattr(result, "cess_amount", ""),
            getattr(result, "total_amount", ""),
            result.confidence_flag,
            source_image or "",
            notes or "",
        ]

        row_num = ws.max_row + 1
        ws.append(row_data)

        # Style the new row
        flag_color = FLAG_COLORS.get(result.confidence_flag, "FFFFFF")
        alt_fill = ALT_ROW_FILL if row_num % 2 == 0 else None
        conf_fill = PatternFill("solid", fgColor=flag_color)
        conf_col_idx = HEADERS.index("Confidence") + 1  # 1-indexed

        for col_idx, cell in enumerate(ws[row_num], start=1):
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col_idx == conf_col_idx:
                cell.fill = conf_fill
            elif alt_fill:
                cell.fill = alt_fill

        ws.row_dimensions[row_num].height = 16

    wb.save(file_path)
    logger.info("Appended %d rows to %s — confidence=%s", 
                len(result.items), file_path.name, result.confidence_flag)
    return file_path
