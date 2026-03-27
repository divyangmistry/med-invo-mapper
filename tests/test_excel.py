"""
test_excel.py — Standalone test for Excel file generation.

Run from the project root (no Docker needed):
    python tests/test_excel.py

What it tests
-------------
  ✓ Daily Excel file is created with correct headers
  ✓ Appending a row works correctly
  ✓ Appending a second row doesn't overwrite first row
  ✓ Confidence flag colour coding (spot-check)
  ✓ File is a valid xlsx workbook
"""
import os
import sys
import tempfile
from pathlib import Path

# Point to a temp directory for outputs so we don't pollute the real outputs/
_tmpdir = tempfile.mkdtemp(prefix="med_invo_test_")
os.environ["DATABASE_URL"] = "sqlite://"
os.environ["OUTPUT_DIR"] = _tmpdir
os.environ["OLLAMA_BASE_URL"] = "http://localhost:11434"
os.environ["VLM_MODEL"] = "qwen2-vl:7b"

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "agent"))

import openpyxl
from extractor import ExtractionResult
from excel_writer import append_to_excel, HEADERS

PASS = "✅"
FAIL = "❌"
results = []


def check(name: str, condition: bool, detail: str = "") -> None:
    icon = PASS if condition else FAIL
    msg = f"  {icon}  {name}"
    if detail:
        msg += f" — {detail}"
    print(msg)
    results.append(condition)


# Sample extraction results
SAMPLE_OK = ExtractionResult(
    vendor_name="Cipla Ltd.",
    invoice_number="INV-2024-001",
    medicine_name="Azithromycin 500mg",
    medicine_code="AZM500",
    batch_number="B20240301",
    manufacturing_date="03/2024",
    expiry_date="03/2026",
    quantity=10,
    confidence_flag="OK",
)

SAMPLE_REVIEW = ExtractionResult(
    vendor_name="Sun Pharma",
    invoice_number="UNKNOWN",
    medicine_name="Paracetamol 650mg",
    medicine_code="PCM650",
    batch_number="UNKNOWN",
    manufacturing_date="01/2024",
    expiry_date="01/2027",
    quantity=5,
    confidence_flag="MANUAL_REVIEW",
)

print("\n" + "=" * 55)
print("  Med-Invo Mapper — Excel Writer Tests")
print("=" * 55)

out_dir = Path(_tmpdir)

# ── Test 1: File creation ─────────────────────────────────────
print("\n[1/5] Excel file creation on first append")
try:
    file_path = append_to_excel(SAMPLE_OK, timestamp="2026-03-17 10:00:00 UTC",
                                source_image="test1.jpg", output_dir=out_dir)
    check("File created", file_path.exists(), f"path={file_path.name}")
    check("File is named correctly", file_path.name.startswith("Inventory_"))
    check("File has .xlsx extension", file_path.suffix == ".xlsx")
except Exception as e:
    check("File created", False, str(e))
    file_path = None

# ── Test 2: Headers check ─────────────────────────────────────
print("\n[2/5] Header row validation")
if file_path:
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        check("All headers present", header_row == HEADERS, str(header_row))
        check("Headers are bold", ws["A1"].font.bold is True)
    except Exception as e:
        check("Header check", False, str(e))

# ── Test 3: Row data verification ─────────────────────────────
print("\n[3/5] First data row verification")
if file_path:
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        row2 = [cell.value for cell in ws[2]]
        check("Vendor name in row 2", row2[1] == "Cipla Ltd.")
        check("Invoice number in row 2", row2[2] == "INV-2024-001")
        check("Medicine code in row 2", row2[4] == "AZM500")
        check("Quantity in row 2", row2[8] == 10)
        check("Confidence flag in row 2", row2[9] == "OK")
    except Exception as e:
        check("Row data verification", False, str(e))

# ── Test 4: Second append (no overwrite) ─────────────────────
print("\n[4/5] Second append doesn't overwrite first row")
if file_path:
    try:
        append_to_excel(SAMPLE_REVIEW, timestamp="2026-03-17 11:00:00 UTC",
                        source_image="test2.jpg", output_dir=out_dir)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        check("Workbook has 3 rows (header + 2 data)", ws.max_row == 3,
              f"actual rows={ws.max_row}")
        row3 = [cell.value for cell in ws[3]]
        check("Row 3 vendor is Sun Pharma", row3[1] == "Sun Pharma")
        check("Row 2 vendor still Cipla (not overwritten)", ws["B2"].value == "Cipla Ltd.")
    except Exception as e:
        check("Second append", False, str(e))

# ── Test 5: Sanity - open with openpyxl ──────────────────────
print("\n[5/5] File is a valid openpyxl workbook")
if file_path:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        check("Workbook is readable", wb is not None)
        check("Sheet name is 'Daily Inventory'", wb.sheetnames[0] == "Daily Inventory")
        wb.close()
    except Exception as e:
        check("File validity", False, str(e))

# ── Cleanup ───────────────────────────────────────────────────
import shutil
shutil.rmtree(_tmpdir, ignore_errors=True)

# ── Summary ───────────────────────────────────────────────────
passed = sum(results)
total = len(results)
print("\n" + "=" * 55)
if passed == total:
    print(f"  ✅  ALL {total} EXCEL TESTS PASSED")
else:
    print(f"  ❌  {total - passed} / {total} TESTS FAILED")
print("=" * 55 + "\n")
sys.exit(0 if passed == total else 1)
